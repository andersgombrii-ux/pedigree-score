from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------
# Normalization (match main.py)
# ----------------------------

_PARENS_SUFFIX_RE = re.compile(r"\s*\([^)]*\)\s*$")


def _normalize_name_for_key(name: str) -> str:
    """
    Normalize a horse name for stable row matching (UPSERT key).

    Must be consistent with focus matching rules:
      - take first line only
      - strip whitespace
      - remove trailing '(...)' suffixes like '(NO)'
      - remove '*' markers
      - collapse internal whitespace
      - casefold
    """
    s = (name or "").splitlines()[0].strip()
    s = s.replace("*", "").strip()
    s = _PARENS_SUFFIX_RE.sub("", s).strip()
    s = " ".join(s.split())
    return s.casefold()


def _safe_ancestor_label(name: str) -> str:
    """
    Produce a stable column suffix for an ancestor token.
    Keep unicode letters (e.g., Grasiös) but normalize whitespace.
    """
    s = name.strip()
    s = " ".join(s.split())
    s = s.replace("\n", " ").replace("\t", " ")
    return s


def _required_headers(focus_ancestors: list[str]) -> list[str]:
    """
    Header order per spec:

      HorseId
      Name
      BirthYear

      Score_Exp_<A...> (per focus ancestor)
      Score_Exp_Total

      Score_SlowExp_<A...> (per focus ancestor)
      Score_SlowExp_Total

      Count_<A...> (per focus ancestor)
    """
    focus = [_safe_ancestor_label(a) for a in focus_ancestors]

    headers: list[str] = ["HorseId", "Name", "BirthYear"]

    headers += [f"Score_Exp_{a}" for a in focus]
    headers += ["Score_Exp_Total"]

    headers += [f"Score_SlowExp_{a}" for a in focus]
    headers += ["Score_SlowExp_Total"]

    headers += [f"Count_{a}" for a in focus]

    return headers


def _get_or_create_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(title=sheet_name)


def _read_headers(ws: Worksheet) -> list[str]:
    if ws.max_row < 1:
        return []
    row = ws[1]
    out: list[str] = []
    for cell in row:
        v = cell.value
        out.append(str(v) if v is not None else "")
    while out and out[-1] == "":
        out.pop()
    return out


def _write_headers(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)


def _ensure_headers(ws: Worksheet, required: list[str]) -> list[str]:
    """
    Ensure headers are consistent across runs.
    If new columns are required, append them and backfill existing rows with 0.
    """
    existing = _read_headers(ws)

    if not existing:
        _write_headers(ws, required)
        return required

    existing_set = set(existing)
    new_cols = [h for h in required if h not in existing_set]

    if not new_cols:
        return existing

    updated = existing + new_cols
    _write_headers(ws, updated)

    if ws.max_row >= 2:
        for r in range(2, ws.max_row + 1):
            for c in range(len(existing) + 1, len(updated) + 1):
                ws.cell(row=r, column=c, value=0)

    return updated


def _cell_int_or_none(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        # Excel may store ints as floats
        if abs(v - int(v)) < 1e-9:
            return int(v)
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        if s.isdigit():
            return int(s)
        return None
    return None


def _cell_str(v: Any) -> str:
    return "" if v is None else str(v)


def _find_matching_rows(
    ws: Worksheet,
    header_to_col: dict[str, int],
    *,
    horse_id: Optional[int],
    horse_name: str,
    birth_year: Optional[int],
) -> list[int]:
    """
    Return worksheet row indices (>=2) that match the UPSERT key.

    Key rules:
      - If HorseId present: HorseId + BirthYear
      - Else: NormalizedName + BirthYear
    """
    by_col_id = header_to_col.get("HorseId")
    by_col_name = header_to_col.get("Name")
    by_col_year = header_to_col.get("BirthYear")

    if by_col_year is None:
        return []

    target_year = birth_year if birth_year is not None else None

    target_norm_name = _normalize_name_for_key(horse_name)

    matches: list[int] = []

    for r in range(2, ws.max_row + 1):
        row_year = _cell_int_or_none(ws.cell(row=r, column=by_col_year).value)

        # BirthYear must match for both key types
        if row_year != target_year:
            continue

        if horse_id is not None and by_col_id is not None:
            row_id = _cell_int_or_none(ws.cell(row=r, column=by_col_id).value)
            if row_id == horse_id:
                matches.append(r)
            continue

        # HorseId missing => fallback key on normalized name + year
        if by_col_name is None:
            continue
        row_name = _cell_str(ws.cell(row=r, column=by_col_name).value)
        if _normalize_name_for_key(row_name) == target_norm_name:
            matches.append(r)

    return matches


def append_scores_row(
    *,
    xlsx_path: Path,
    sheet_name: str,
    horse_id: Optional[int],
    horse_name: str,
    birth_year: Optional[int],
    focus_ancestors: list[str],
    per_ancestor: dict[str, dict[str, float]],
) -> None:
    """
    UPSERT exactly one row into the Excel table.

    Behavior:
      - If file doesn't exist: create with headers.
      - If matching row exists (stable key):
          - overwrite that row (replace duplicates: keep first, delete extras)
      - If no match: append a new row.
      - If focus ancestor list changes: add missing columns and backfill old rows with 0.
      - Missing ancestor values -> 0.

    Stable key:
      - If HorseId present: (HorseId, BirthYear)
      - Else: (NormalizedName(Name), BirthYear)

    per_ancestor format:
      {
        "<token>": {"score_exp": float, "score_exp_slow": float, "count": float},
        ...
      }
    """
    xlsx_path = Path(xlsx_path)
    required = _required_headers(focus_ancestors)

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    ws = _get_or_create_sheet(wb, sheet_name)

    # Remove default "Sheet" if it's empty and we're creating a new named sheet
    if xlsx_path.exists() is False and "Sheet" in wb.sheetnames and sheet_name != "Sheet":
        default_ws = wb["Sheet"]
        if default_ws.max_row == 1 and default_ws.max_column == 1 and default_ws["A1"].value is None:
            wb.remove(default_ws)

    headers = _ensure_headers(ws, required)
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    # Base values
    row_data: dict[str, Any] = {
        "HorseId": horse_id if horse_id is not None else "",
        "Name": horse_name,
        "BirthYear": birth_year if birth_year is not None else "",
    }

    # Per-ancestor values + totals (missing treated as 0)
    exp_total = 0.0
    slow_total = 0.0

    for tok in focus_ancestors:
        label = _safe_ancestor_label(tok)
        d = per_ancestor.get(tok) or {}

        exp = float(d.get("score_exp", 0.0) or 0.0)
        slow = float(d.get("score_exp_slow", 0.0) or 0.0)
        cnt = float(d.get("count", 0.0) or 0.0)

        row_data[f"Score_Exp_{label}"] = exp
        row_data[f"Score_SlowExp_{label}"] = slow
        row_data[f"Count_{label}"] = int(cnt) if abs(cnt - int(cnt)) < 1e-9 else cnt

        exp_total += exp
        slow_total += slow

    row_data["Score_Exp_Total"] = exp_total
    row_data["Score_SlowExp_Total"] = slow_total

    # Fill missing columns with 0 where appropriate
    for h in headers:
        if h not in row_data:
            if h.startswith("Score_") or h.startswith("Count_"):
                row_data[h] = 0
            else:
                row_data[h] = ""

    # --- UPSERT ---
    matching_rows = _find_matching_rows(
        ws,
        header_to_col,
        horse_id=horse_id,
        horse_name=horse_name,
        birth_year=birth_year,
    )

    if matching_rows:
        target_row = matching_rows[0]

        # Overwrite values in target row
        for h, v in row_data.items():
            col = header_to_col.get(h)
            if col is None:
                continue
            ws.cell(row=target_row, column=col, value=v)

        # Delete any extra duplicates (from bottom to top to preserve indices)
        if len(matching_rows) > 1:
            for r in sorted(matching_rows[1:], reverse=True):
                ws.delete_rows(r, 1)

    else:
        # Append new row
        new_row_idx = ws.max_row + 1 if ws.max_row >= 1 else 2
        for h, v in row_data.items():
            col = header_to_col.get(h)
            if col is None:
                continue
            ws.cell(row=new_row_idx, column=col, value=v)

    wb.save(xlsx_path)