from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.scores_xlsx import append_scores_row


def _read_rows(path: Path, sheet: str) -> list[dict[str, object]]:
    wb = load_workbook(path)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        out.append({headers[i]: ws.cell(row=r, column=i + 1).value for i in range(len(headers))})
    return out


def test_scores_xlsx_upsert_by_horseid_birthyear(tmp_path: Path) -> None:
    xlsx = tmp_path / "scores.xlsx"
    sheet = "Scores"
    focus = ["Dalterna", "Kaprell"]

    # First write
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=491586,
        horse_name="VOJE LITA (NO)",
        birth_year=1983,
        focus_ancestors=focus,
        per_ancestor={
            "Dalterna": {"score_exp": 0.1, "score_exp_slow": 0.2, "count": 5.0},
            "Kaprell": {"score_exp": 0.0, "score_exp_slow": 0.0, "count": 0.0},
        },
    )

    rows1 = _read_rows(xlsx, sheet)
    assert len(rows1) == 1
    assert rows1[0]["HorseId"] == 491586
    assert rows1[0]["BirthYear"] == 1983

    # Second write for same key, changed values => must overwrite, not append
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=491586,
        horse_name="VOJE LITA (NO)",
        birth_year=1983,
        focus_ancestors=focus,
        per_ancestor={
            "Dalterna": {"score_exp": 0.1, "score_exp_slow": 0.2, "count": 5.0},
            "Kaprell": {"score_exp": 0.25, "score_exp_slow": 0.5625, "count": 1.0},
        },
    )

    rows2 = _read_rows(xlsx, sheet)
    assert len(rows2) == 1  # UPSERT: row count unchanged
    assert rows2[0]["Count_Kaprell"] == 1
    assert abs(float(rows2[0]["Score_Exp_Kaprell"]) - 0.25) < 1e-12


def test_scores_xlsx_upsert_fallback_normalizedname_birthyear(tmp_path: Path) -> None:
    xlsx = tmp_path / "scores.xlsx"
    sheet = "Scores"
    focus = ["Dalterna"]

    # HorseId missing => fallback to normalized name + birthyear
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=None,
        horse_name="VÅRBLOMSTER",
        birth_year=2021,
        focus_ancestors=focus,
        per_ancestor={"Dalterna": {"score_exp": 0.0, "score_exp_slow": 0.0, "count": 1.0}},
    )

    # Same horse, name formatted differently should still match
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=None,
        horse_name="Vårblomster* (SE)",
        birth_year=2021,
        focus_ancestors=focus,
        per_ancestor={"Dalterna": {"score_exp": 0.0, "score_exp_slow": 0.0, "count": 2.0}},
    )

    rows = _read_rows(xlsx, sheet)
    assert len(rows) == 1
    assert rows[0]["BirthYear"] == 2021
    assert rows[0]["Count_Dalterna"] == 2


def test_scores_xlsx_upsert_removes_duplicates(tmp_path: Path) -> None:
    xlsx = tmp_path / "scores.xlsx"
    sheet = "Scores"
    focus = ["Dalterna"]

    # Create two duplicate rows by appending with different HorseId (missing) but same fallback key
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=None,
        horse_name="TESTHORSE (NO)",
        birth_year=2000,
        focus_ancestors=focus,
        per_ancestor={"Dalterna": {"score_exp": 0.0, "score_exp_slow": 0.0, "count": 1.0}},
    )
    # Force a second duplicate (simulating old behavior)
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=None,
        horse_name="TESTHORSE",
        birth_year=2000,
        focus_ancestors=focus,
        per_ancestor={"Dalterna": {"score_exp": 0.0, "score_exp_slow": 0.0, "count": 1.0}},
    )

    # Now upsert again: should collapse duplicates to single row
    append_scores_row(
        xlsx_path=xlsx,
        sheet_name=sheet,
        horse_id=None,
        horse_name="Testhorse* (NO)",
        birth_year=2000,
        focus_ancestors=focus,
        per_ancestor={"Dalterna": {"score_exp": 0.0, "score_exp_slow": 0.0, "count": 3.0}},
    )

    rows = _read_rows(xlsx, sheet)
    assert len(rows) == 1
    assert rows[0]["Count_Dalterna"] == 3